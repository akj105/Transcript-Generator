import csv
import openpyxl
import os
os.system('cls')
from pywebio.input import *
from pywebio.output import *
from pywebio import start_server
from reportlab.lib.pagesizes import A3
from reportlab.platypus import Paragraph, Table
from reportlab.platypus import TableStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet
import datetime

def generate_marksheet(grade):
    path=r"./output/"
    colm=("Subject no.","Subject Name","L-T-P","Credit","Grade")
    a=1
    b=0
    for line in grade:
        roll= line[0]
        sem = line[1]
        if roll!="Roll":
            if os.path.exists(path+'%s.xlsx' %roll)== False:
                wb = openpyxl.Workbook()
                sheet=wb.active
                sheet['A1']="Roll No."
                sheet['B1']=roll
                sheet['A2']="Name of Student"
                sheet['B2']=dict1[roll]
                sheet['A3']="Discipline"
                sheet['B3']=roll[4:6]
                sheet['A4']="Semester No."
                sheet['A5']="Semester wise Credit Taken"
                sheet['A6']="SPI"
                sheet['A7']="Total Credits Taken"
                sheet['A8']="CPI"
                sheet.title="Overall"
                wb.save(path+'%s.xlsx' %roll)
                wb=openpyxl.load_workbook(path+'%s.xlsx' %roll)
                wb.create_sheet(index=1, title="Sem1")
                b=1
                sheet=wb["Sem1"]
                a=1
                sheet.append(colm)
                details=(line[2],dict2[line[2]],dict3[line[2]],line[3],line[4])
                sheet.append(details)
                a+=1
                wb.save(path+'%s.xlsx' %roll)
            else:
                wb=openpyxl.load_workbook(path+'%s.xlsx' %roll)
                if(b!=int(line[1])):
                    b+=1
                    sheetname="Sem"+str(b)
                    wb.create_sheet(index=b,title=sheetname)
                    sheet=wb[sheetname]
                    sheet.append(colm)
                    a=1
                currsheet="Sem"+str(b)
                sheet=wb[currsheet]
                details=(line[2],dict2[line[2]],dict3[line[2]],line[3],line[4])
                sheet.append(details)
                a+=1
                wb.save(path+'%s.xlsx' %roll)

    return


def generate_marksheet_range(grade):
    path=r"./output/"
    colm=("Subject no.","Subject Name","L-T-P","Credit","Grade")
    a=1
    b=0
    for line in grade:
        roll= line[0]
        sem = line[1]
        if roll in roll_list:
            if os.path.exists(path+'%s.xlsx' %roll)== False:
                wb = openpyxl.Workbook()
                sheet=wb.active
                sheet['A1']="Roll No."
                sheet['B1']=roll
                sheet['A2']="Name of Student"
                sheet['B2']=dict1[roll]
                sheet['A3']="Discipline"
                sheet['B3']=roll[4:6]
                sheet['A4']="Semester No."
                sheet['A5']="Semester wise Credit Taken"
                sheet['A6']="SPI"
                sheet['A7']="Total Credits Taken"
                sheet['A8']="CPI"
                sheet.title="Overall"
                wb.save(path+'%s.xlsx' %roll)
                wb=openpyxl.load_workbook(path+'%s.xlsx' %roll)
                wb.create_sheet(index=1, title="Sem1")
                b=1
                sheet=wb["Sem1"]
                a=1
                sheet.append(colm)
                details=(line[2],dict2[line[2]],dict3[line[2]],line[3],line[4])
                sheet.append(details)
                a+=1
                wb.save(path+'%s.xlsx' %roll)
            else:
                wb=openpyxl.load_workbook(path+'%s.xlsx' %roll)
                if(b!=int(line[1])):
                    b+=1
                    sheetname="Sem"+str(b)
                    wb.create_sheet(index=b,title=sheetname)
                    sheet=wb[sheetname]
                    sheet.append(colm)
                    a=1
                currsheet="Sem"+str(b)
                sheet=wb[currsheet]
                details=(line[2],dict2[line[2]],dict3[line[2]],line[3],line[4])
                sheet.append(details)
                a+=1
                wb.save(path+'%s.xlsx' %roll)

    return


def cpi_cal(grades):
    path=r"./output/"
    a=0
    b=0
    c=0
    d=0
    e=0
    curr_roll="Roll"
    for line in grades:
        roll = line[0]
        if roll!="Roll":
            if curr_roll!=roll:
                curr_roll=roll
                a=0
                b=0
                c=0
                d=0
                e=0
            wb=openpyxl.load_workbook(path+'%s.xlsx' %roll)
            sheet=wb.active
            if(b!=int(line[1])):
               b+=1
               a=0
               c=0
            a+=int(line[3])
            d+=int(line[3])
            c+=(int(line[3])*dict4[line[4].strip()])
            e+=(int(line[3])*dict4[line[4].strip()])
            sheet.cell(row=4,column=b+1).value=b
            sheet.cell(row=5,column=b+1).value=a
            sheet.cell(row=6,column=b+1).value=round(c/a,2)
            sheet.cell(row=7,column=b+1).value=d
            sheet.cell(row=8,column=b+1).value=round(e/d,2)

            wb.save(path+'%s.xlsx' %roll)
            
    return

def cpi_cal_range(grades):
    path=r"./output/"
    a=0
    b=0
    c=0
    d=0
    e=0
    curr_roll="Roll"
    for line in grades:
        roll = line[0]
        if roll in roll_list:
            if curr_roll!=roll:
                curr_roll=roll
                a=0
                b=0
                c=0
                d=0
                e=0
            wb=openpyxl.load_workbook(path+'%s.xlsx' %roll)
            sheet=wb.active
            if(b!=int(line[1])):
               b+=1
               a=0
               c=0
            a+=int(line[3])
            d+=int(line[3])
            c+=(int(line[3])*dict4[line[4].strip()])
            e+=(int(line[3])*dict4[line[4].strip()])
            sheet.cell(row=4,column=b+1).value=b
            sheet.cell(row=5,column=b+1).value=a
            sheet.cell(row=6,column=b+1).value=round(c/a,2)
            sheet.cell(row=7,column=b+1).value=d
            sheet.cell(row=8,column=b+1).value=round(e/d,2)

            wb.save(path+'%s.xlsx' %roll)
            
    return



def generate_pdf(line):
    roll=line
    path=r"./output/"
    if roll!="Roll":
        filename=r"./transcriptsIITP/"+'%s.pdf' %roll
        pdf = canvas.Canvas(filename, pagesize=A3)
        width,height=A3
        wb=openpyxl.load_workbook(path+'%s.xlsx' %roll)

        p0text="Roll no:"+roll +", Name: "+dict1[roll] +", Year of admission: "+str(20)+roll[0:2]+", Programme: "+dict5[roll[2:4]]+", Course: "+roll[4:6]  
        styles = getSampleStyleSheet()
        p0 = Paragraph(p0text, style=styles["Normal"])
        p0.wrapOn(pdf, 280*mm, 10*mm)  
        p0.drawOn(pdf, 50*mm, 385*mm)
        pdf.drawInlineImage("pic1.png", 10*mm, 395*mm)

        current_time = datetime.datetime.now()
        p2text="Date Generated: "+str(current_time.day) +"/"+str(current_time.month)+"/"+str(current_time.year)+", Time: "+str(current_time.hour)+":"+str(current_time.minute)+":"+str(current_time.second)  
        p2 = Paragraph(p2text, style=styles["Normal"])
        p2.wrapOn(pdf, 100*mm, 10*mm)  
        p2.drawOn(pdf, 10*mm, 50*mm)   

        p3text="Assitant Registrar (Academic)"
        p3 = Paragraph(p3text, style=styles["Normal"])
        p3.wrapOn(pdf, 100*mm, 10*mm)  
        p3.drawOn(pdf, 230*mm, 40*mm)   

        if seal_option == "YES":
            pdf.drawInlineImage("seal.jpeg", 110*mm, 30*mm) 
            pdf.drawInlineImage("sign.jpeg", 220*mm, 50*mm) 
             

        if "Sem1" in wb.sheetnames:
            currsheet="Sem1"
            sheet=wb[currsheet]
            data=[
                [sheet.cell(row=1,column=1).value,sheet.cell(row=1,column=2).value,sheet.cell(row=1,column=3).value,sheet.cell(row=1,column=4).value,sheet.cell(row=1,column=5).value]   
            ]
            i=1
            while(sheet.cell(row=i+1,column=1).value):
                data.append([sheet.cell(row=i+1,column=1).value,sheet.cell(row=i+1,column=2).value,sheet.cell(row=i+1,column=3).value,sheet.cell(row=i+1,column=4).value,sheet.cell(row=i+1,column=5).value])
                i=i+1

            table = Table(data,rowHeights=5*mm)
            style= TableStyle([
                ('GRID', (0,0),(-1,-1),1,colors.black)
            ])
            table.setStyle(style)
            table.wrapOn(pdf, width, height)
            table.drawOn(pdf, 10*mm, 335*mm)

            sheet=wb["Overall"]
            p1text = "SEMESTER 1 : Credits taken:"+ str(sheet.cell(row=5,column=2).value)+ ", Credits cleared:"+ str(sheet.cell(row=5,column=2).value)+", SPI:"+str(sheet.cell(row=6,column=2).value)+ ", CPI:"+str(sheet.cell(row=8,column=2).value) 
            p1 = Paragraph(p1text, style=styles["Normal"])
            p1.wrapOn(pdf, 120*mm, 10*mm)  
            p1.drawOn(pdf, 10*mm, 325*mm)  

        else:
            wb.save(path+'%s.xlsx' %roll)
            pdf.save()
            return

        if "Sem2" in wb.sheetnames:
            currsheet="Sem2"
            sheet=wb[currsheet]
            data=[
                [sheet.cell(row=1,column=1).value,sheet.cell(row=1,column=2).value,sheet.cell(row=1,column=3).value,sheet.cell(row=1,column=4).value,sheet.cell(row=1,column=5).value]   
            ]
            i=1
            while(sheet.cell(row=i+1,column=1).value):
                data.append([sheet.cell(row=i+1,column=1).value,sheet.cell(row=i+1,column=2).value,sheet.cell(row=i+1,column=3).value,sheet.cell(row=i+1,column=4).value,sheet.cell(row=i+1,column=5).value])
                i=i+1

            table = Table(data,rowHeights=5*mm)
            style= TableStyle([
                ('GRID', (0,0),(-1,-1),1,colors.black)
            ])
            table.setStyle(style)
            table.wrapOn(pdf, width, height)
            table.drawOn(pdf, 160*mm, 335*mm)

            sheet=wb["Overall"]
            p1text = "SEMESTER 2 : Credits taken:"+ str(sheet.cell(row=5,column=3).value)+ ", Credits cleared:"+ str(sheet.cell(row=5,column=3).value)+", SPI:"+str(sheet.cell(row=6,column=3).value)+ ", CPI:"+str(sheet.cell(row=8,column=3).value) 
            p1 = Paragraph(p1text, style=styles["Normal"])
            p1.wrapOn(pdf, 120*mm, 10*mm)  
            p1.drawOn(pdf, 160*mm, 325*mm)  

        else:
            wb.save(path+'%s.xlsx' %roll)
            pdf.save()
            return

        if "Sem3" in wb.sheetnames:
            currsheet="Sem3"
            sheet=wb[currsheet]
            data=[
                [sheet.cell(row=1,column=1).value,sheet.cell(row=1,column=2).value,sheet.cell(row=1,column=3).value,sheet.cell(row=1,column=4).value,sheet.cell(row=1,column=5).value]   
            ]
            i=1
            while(sheet.cell(row=i+1,column=1).value):
                data.append([sheet.cell(row=i+1,column=1).value,sheet.cell(row=i+1,column=2).value,sheet.cell(row=i+1,column=3).value,sheet.cell(row=i+1,column=4).value,sheet.cell(row=i+1,column=5).value])
                i=i+1

            table = Table(data,rowHeights=5*mm)
            style= TableStyle([
                ('GRID', (0,0),(-1,-1),1,colors.black)
            ])
            table.setStyle(style)
            table.wrapOn(pdf, width, height)
            table.drawOn(pdf, 10*mm, 270*mm)

            sheet=wb["Overall"]
            p1text = "SEMESTER 3 : Credits taken:"+ str(sheet.cell(row=5,column=4).value)+ ", Credits cleared:"+ str(sheet.cell(row=5,column=4).value)+", SPI:"+str(sheet.cell(row=6,column=4).value)+ ", CPI:"+str(sheet.cell(row=8,column=4).value) 
            p1 = Paragraph(p1text, style=styles["Normal"])
            p1.wrapOn(pdf, 120*mm, 10*mm)  
            p1.drawOn(pdf, 10*mm, 260*mm)  

        else:
            wb.save(path+'%s.xlsx' %roll)
            pdf.save()
            return


        if "Sem4" in wb.sheetnames:
            currsheet="Sem4"
            sheet=wb[currsheet]
            data=[
                [sheet.cell(row=1,column=1).value,sheet.cell(row=1,column=2).value,sheet.cell(row=1,column=3).value,sheet.cell(row=1,column=4).value,sheet.cell(row=1,column=5).value]   
            ]
            i=1
            while(sheet.cell(row=i+1,column=1).value):
                data.append([sheet.cell(row=i+1,column=1).value,sheet.cell(row=i+1,column=2).value,sheet.cell(row=i+1,column=3).value,sheet.cell(row=i+1,column=4).value,sheet.cell(row=i+1,column=5).value])
                i=i+1

            table = Table(data,rowHeights=5*mm)
            style= TableStyle([
                ('GRID', (0,0),(-1,-1),1,colors.black)
            ])
            table.setStyle(style)
            table.wrapOn(pdf, width, height)
            table.drawOn(pdf, 160*mm, 270*mm)

            sheet=wb["Overall"]
            p1text = "SEMESTER 4 : Credits taken:"+ str(sheet.cell(row=5,column=5).value)+ ", Credits cleared:"+ str(sheet.cell(row=5,column=5).value)+", SPI:"+str(sheet.cell(row=6,column=5).value)+ ", CPI:"+str(sheet.cell(row=8,column=5).value) 
            p1 = Paragraph(p1text, style=styles["Normal"])
            p1.wrapOn(pdf, 120*mm, 10*mm)  
            p1.drawOn(pdf, 160*mm, 260*mm)  

        else:
            wb.save(path+'%s.xlsx' %roll)
            pdf.save()
            return
        
        if "Sem5" in wb.sheetnames:
            currsheet="Sem5"
            sheet=wb[currsheet]
            data=[
                [sheet.cell(row=1,column=1).value,sheet.cell(row=1,column=2).value,sheet.cell(row=1,column=3).value,sheet.cell(row=1,column=4).value,sheet.cell(row=1,column=5).value]   
            ]
            i=1
            while(sheet.cell(row=i+1,column=1).value):
                data.append([sheet.cell(row=i+1,column=1).value,sheet.cell(row=i+1,column=2).value,sheet.cell(row=i+1,column=3).value,sheet.cell(row=i+1,column=4).value,sheet.cell(row=i+1,column=5).value])
                i=i+1

            table = Table(data,rowHeights=5*mm)
            style= TableStyle([
                ('GRID', (0,0),(-1,-1),1,colors.black)
            ])
            table.setStyle(style)
            table.wrapOn(pdf, width, height)
            table.drawOn(pdf, 10*mm, 205*mm)

            sheet=wb["Overall"]
            p1text = "SEMESTER 5 : Credits taken:"+ str(sheet.cell(row=5,column=6).value)+ ", Credits cleared:"+ str(sheet.cell(row=5,column=6).value)+", SPI:"+str(sheet.cell(row=6,column=6).value)+ ", CPI:"+str(sheet.cell(row=8,column=6).value) 
            p1 = Paragraph(p1text, style=styles["Normal"])
            p1.wrapOn(pdf, 120*mm, 10*mm)  
            p1.drawOn(pdf, 10*mm, 195*mm)  

        else:
            wb.save(path+'%s.xlsx' %roll)
            pdf.save()
            return

        if "Sem6" in wb.sheetnames:
            currsheet="Sem6"
            sheet=wb[currsheet]
            data=[
                [sheet.cell(row=1,column=1).value,sheet.cell(row=1,column=2).value,sheet.cell(row=1,column=3).value,sheet.cell(row=1,column=4).value,sheet.cell(row=1,column=5).value]   
            ]
            i=1
            while(sheet.cell(row=i+1,column=1).value):
                data.append([sheet.cell(row=i+1,column=1).value,sheet.cell(row=i+1,column=2).value,sheet.cell(row=i+1,column=3).value,sheet.cell(row=i+1,column=4).value,sheet.cell(row=i+1,column=5).value])
                i=i+1

            table = Table(data,rowHeights=5*mm)
            style= TableStyle([
                ('GRID', (0,0),(-1,-1),1,colors.black)
            ])
            table.setStyle(style)
            table.wrapOn(pdf, width, height)
            table.drawOn(pdf, 160*mm, 205*mm)

            sheet=wb["Overall"]
            p1text = "SEMESTER 6 : Credits taken:"+ str(sheet.cell(row=5,column=7).value)+ ", Credits cleared:"+ str(sheet.cell(row=5,column=7).value)+", SPI:"+str(sheet.cell(row=6,column=7).value)+ ", CPI:"+str(sheet.cell(row=8,column=7).value) 
            p1 = Paragraph(p1text, style=styles["Normal"])
            p1.wrapOn(pdf, 120*mm, 10*mm)  
            p1.drawOn(pdf, 160*mm, 195*mm)  

        else:
            wb.save(path+'%s.xlsx' %roll)
            pdf.save()
            return

        if "Sem7" in wb.sheetnames:
            currsheet="Sem7"
            sheet=wb[currsheet]
            data=[
                [sheet.cell(row=1,column=1).value,sheet.cell(row=1,column=2).value,sheet.cell(row=1,column=3).value,sheet.cell(row=1,column=4).value,sheet.cell(row=1,column=5).value]   
            ]
            i=1
            while(sheet.cell(row=i+1,column=1).value):
                data.append([sheet.cell(row=i+1,column=1).value,sheet.cell(row=i+1,column=2).value,sheet.cell(row=i+1,column=3).value,sheet.cell(row=i+1,column=4).value,sheet.cell(row=i+1,column=5).value])
                i=i+1

            table = Table(data,rowHeights=5*mm)
            style= TableStyle([
                ('GRID', (0,0),(-1,-1),1,colors.black)
            ])
            table.setStyle(style)
            table.wrapOn(pdf, width, height)
            table.drawOn(pdf, 10*mm, 140*mm)

            sheet=wb["Overall"]
            p1text = "SEMESTER 7 : Credits taken:"+ str(sheet.cell(row=5,column=8).value)+ ", Credits cleared:"+ str(sheet.cell(row=5,column=8).value)+", SPI:"+str(sheet.cell(row=6,column=8).value)+ ", CPI:"+str(sheet.cell(row=8,column=8).value) 
            p1 = Paragraph(p1text, style=styles["Normal"])
            p1.wrapOn(pdf, 120*mm, 10*mm)  
            p1.drawOn(pdf, 10*mm, 130*mm)  

        else:
            wb.save(path+'%s.xlsx' %roll)
            pdf.save()
            return

        if "Sem8" in wb.sheetnames:
            currsheet="Sem8"
            sheet=wb[currsheet]
            data=[
                [sheet.cell(row=1,column=1).value,sheet.cell(row=1,column=2).value,sheet.cell(row=1,column=3).value,sheet.cell(row=1,column=4).value,sheet.cell(row=1,column=5).value]   
            ]
            i=1
            while(sheet.cell(row=i+1,column=1).value):
                data.append([sheet.cell(row=i+1,column=1).value,sheet.cell(row=i+1,column=2).value,sheet.cell(row=i+1,column=3).value,sheet.cell(row=i+1,column=4).value,sheet.cell(row=i+1,column=5).value])
                i=i+1

            table = Table(data,rowHeights=5*mm)
            style= TableStyle([
                ('GRID', (0,0),(-1,-1),1,colors.black)
            ])
            table.setStyle(style)
            table.wrapOn(pdf, width, height)
            table.drawOn(pdf, 160*mm, 140*mm)

            sheet=wb["Overall"]
            p1text = "SEMESTER 8 : Credits taken:"+ str(sheet.cell(row=5,column=9).value)+ ", Credits cleared:"+ str(sheet.cell(row=5,column=9).value)+", SPI:"+str(sheet.cell(row=6,column=9).value)+ ", CPI:"+str(sheet.cell(row=8,column=9).value) 
            p1 = Paragraph(p1text, style=styles["Normal"])
            p1.wrapOn(pdf, 120*mm, 10*mm)  
            p1.drawOn(pdf, 160*mm, 130*mm)  

        else:
            wb.save(path+'%s.xlsx' %roll)
            pdf.save()
            return



        wb.save(path+'%s.xlsx' %roll)
        pdf.save()

    return



grades = file_upload("Select grades file", accept=".csv")
names_roll = file_upload("Select names-roll file", accept=".csv")
subjects_master = file_upload("Select subjects_master file", accept=".csv")

put_text('Option 1: A range of roll numbers \n')
put_text('Option 2: all transcripts \n')
option = select('Select option:', [1, 2])

roll_list=[]

if option == 1:
    first_roll=input('Enter 1st roll number', type = "text")
    last_roll=input('Enter last roll number', type = "text")
    first_roll=first_roll.upper()
    last_roll=last_roll.upper()
    #put_text('Your range is ', first_roll, ' - ', last_roll)
    roll_list.append(first_roll)
    while(first_roll!=last_roll):
        if first_roll[6] == "0" and first_roll[7] != "9":
            first_roll=str(first_roll[0:7])+str(int(first_roll[7:])+1)
        else:
            first_roll=str(first_roll[0:6])+str(int(first_roll[6:])+1)
        roll_list.append(first_roll)

put_text('Do you want to upload SEAL and signature? \n')
seal_option = select('Select option:', ["YES", "NO"])

if seal_option == "YES":
    seal = file_upload("Upload SEAL", accept="image/*")
    sign = file_upload("Upload signature", accept="image/*")

    with open('seal.jpeg', 'wb') as file:
	    file.write(seal['content'])
if seal_option == "YES":
    with open('sign.jpeg', 'wb') as file:
	    file.write(sign['content'])

content_2 = names_roll['content'].decode('utf-8').splitlines()
rdr= csv.reader(content_2)
title1=next(rdr)
dict1={}
for arr in rdr:
    dict1[arr[0]]=arr[1]


content_3 = subjects_master['content'].decode('utf-8').splitlines()
rd=csv.reader(content_3)
title2=next(rd)
dict2={}
dict3={}
for arr in rd:
    dict2[arr[0]]=arr[1]
    dict3[arr[0]]=arr[2]

dict4={
    "AA":10,
    "AA*":10,
    "AB":9,
    "AB*":9,
    "BB":8,
    "BB*":8,
    "BC":7,
    "BC*":7,
    "CC":6,
    "CC*":6,
    "CD":5,
    "CD*":5,
    "DD":4,
    "DD*":4,
    "F":0,
    "F*":0,
    "I":0,
    "I*":0,
    }

dict5={
    "01":"B.Tech",
    "11":"M.Tech",
    "12":"M.sc",
    "21":"Phd"
}

main_dir="./output"
if os.path.exists(main_dir) == False:
    os.mkdir(main_dir)

main_dir="./transcriptsIITP"
if os.path.exists(main_dir) == False:
    os.mkdir(main_dir)

if option==2:
    content_1 = grades['content'].decode('utf-8').splitlines()
    csv_read=csv.reader(content_1)
    generate_marksheet(csv_read)

    content_1 = grades['content'].decode('utf-8').splitlines()
    c_r=csv.reader(content_1)
    cpi_cal(c_r)

    content_2 = names_roll['content'].decode('utf-8').splitlines()
    c_rr=csv.reader(content_2)
    for arr in c_rr:
        generate_pdf(arr[0])

else:
    content_1 = grades['content'].decode('utf-8').splitlines()
    csv_read=csv.reader(content_1)
    generate_marksheet_range(csv_read)

    content_1 = grades['content'].decode('utf-8').splitlines()
    c_r=csv.reader(content_1)
    cpi_cal_range(c_r)

    for arr in roll_list:
        if arr in dict1:
            generate_pdf(arr)
        else:
            put_text(arr," don't exist.")